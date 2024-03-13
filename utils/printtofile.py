import csv
import os
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

redFill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
darkRedText = Font(color="FF9C0006", name="Calibri", size=10)

greenFill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
darkGreenText = Font(color="FF006100", name="Calibri", size=10)

yellowFill = PatternFill(
    start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"
)
darkYellowText = Font(color="FF9C6500", name="Calibri", size=10)

headerFill = PatternFill("solid", fgColor="47402D")
headerfont = Font(bold=True, color="FFFEFB")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class PrintToFile:
    """Creating files with different types"""

    @staticmethod
    def to_csv(list_of_mo: list, file_to_save: str, csv_header: list) -> bool:
        """Create file to csv"""
        with open(file_to_save, "w", encoding="UTF8", newline="") as f:
            writer = csv.writer(f)

            writer.writerow(csv_header)

            for item in list_of_mo:
                writer.writerow(item)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt(contents: str, file_to_save: str) -> bool:
        """Create file to txt"""
        with open(file_to_save, "w") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt_unix(contents: str, file_to_save: str) -> bool:
        """Create file to txt unix"""
        with open(file_to_save, "w", newline="\n") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_txt_unix_append(contents: str, file_to_save: str) -> bool:
        """Create file to txt unix but append the file"""
        with open(file_to_save, "a", newline="\n") as f:
            f.write(contents)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
    ):
        """Create file to xlsx"""

        # check if workbook is exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create cell
        ws_target = wb.create_sheet(ws_name)

        # starting row
        starting_row = 1

        # header
        for index, header in enumerate(list_of_header):
            ws_target.cell(row=starting_row, column=index + 1).value = header

        starting_row += 1
        for items in list_of_contents:
            for col, item in enumerate(items):
                ws_target.cell(row=starting_row, column=col + 1).value = item

            starting_row += 1

        wb.save(file_to_save)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx_undefined_filled(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
        list_of_red: list,
        list_of_green: list,
    ):
        # check if workbook exists
        if os.path.exists(file_to_save):
            wb = load_workbook(file_to_save)
        else:
            wb = Workbook()
            wb.remove(wb["Sheet"])

        # create or get worksheet
        if ws_name in wb.sheetnames:
            ws_target = wb[ws_name]
            wb.remove(ws_target)
            ws_target = wb.create_sheet(ws_name)
        else:
            ws_target = wb.create_sheet(ws_name)

        # Sort and remove duplicates
        list_of_contents = sorted(list_of_contents, key=lambda x: x[0])
        unique_contents = []
        [
            unique_contents.append(item)
            for item in list_of_contents
            if item not in unique_contents
        ]
        # Freeze the first row
        ws_target.freeze_panes = "A2"

        # header
        for index, header in enumerate(list_of_header):
            ws_target.cell(row=1, column=index + 1).value = header
            ws_target.cell(row=1, column=index + 1).font = Font(
                name="Calibri", size=10, bold=True, color="FFFEFB"
            )
            ws_target.cell(row=1, column=index + 1).fill = headerFill
            ws_target.cell(row=1, column=index + 1).border = thin_border
            ws_target.cell(row=1, column=index + 1).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for row_index, items in enumerate(unique_contents, start=2):
            for col_index, item in enumerate(items, start=1):
                cell = ws_target.cell(row=row_index, column=col_index, value=item)
                if list_of_red and any(ext in str(item) for ext in list_of_red):
                    cell.fill = redFill
                    cell.font = Font(name="Calibri", size=10, color="FF9C0006")
                elif list_of_green and any(ext in str(item) for ext in list_of_green):
                    cell.fill = greenFill
                    cell.font = Font(name="Calibri", size=10, color="FF006100")
                else:
                    cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
        ws_target.auto_filter.ref = f"A1:{get_column_letter(len(list_of_header))}1"
        wb.save(file_to_save)

        return os.path.exists(file_to_save)

    @staticmethod
    def to_xlsx_filled_and_col_offside(
        file_to_save: str,
        ws_name: str,
        list_of_header: list,
        list_of_contents: list,
        list_of_red: list = [],
        col_offside: int = 0,
    ):
        try:
            # check if workbook is exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # create cell
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            # starting row
            starting_row = 1

            # header
            for index, header in enumerate(list_of_header):
                col = index + 1 + col_offside
                ws_target.cell(row=starting_row, column=col).value = header

            starting_row += 1
            for items in list_of_contents:
                for col, item in enumerate(items):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).value = item
                    if item in list_of_red:
                        ws_target.cell(
                            row=starting_row, column=col + 1 + col_offside
                        ).fill = redFill

                starting_row += 1

            wb.save(file_to_save)

            return True

        except Exception:
            errors = traceback.format_exc()
            print(errors)
            return False

    @staticmethod
    def to_xlsx_offside_noheader(
        file_to_save: str,
        ws_name: str,
        list_of_contents: list,
        list_of_red: list = [],
        col_offside: int = 0,
    ):
        try:
            # check if workbook is exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # create cell
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            # starting row
            starting_row = 2

            starting_row += 1
            for items in list_of_contents:
                for col, item in enumerate(items):
                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).value = item
                    if item in list_of_red:
                        ws_target.cell(
                            row=starting_row, column=col + 1 + col_offside
                        ).fill = redFill

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).border = thin_border

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).alignment = Alignment(horizontal="center", vertical="center")

                    ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    ).font = Font(name="Calibri", size=12)

                starting_row += 1

            wb.save(file_to_save)

            return True

        except Exception:
            errors = traceback.format_exc()
            print(errors)
            return False

    @staticmethod
    def to_xlsx_offside_noheader_fullcolor(
        file_to_save: str,
        ws_name: str,
        list_of_contents: list,
        list_of_red: list = [],
        list_of_green: list = [],
        list_of_yellow: list = [],
        col_offside: int = 0,
    ):
        try:
            # Check if workbook exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # Create sheet
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            # Starting row
            starting_row = 4

            starting_row += 1
            for items in list_of_contents:
                for col, item in enumerate(items):
                    cell = ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    )
                    cell.value = item

                    # Apply formatting based on category
                    if item in list_of_red:
                        cell.fill = redFill
                        cell.font = darkRedText

                    elif item in list_of_green:
                        cell.fill = greenFill
                        cell.font = darkGreenText

                    elif item in list_of_yellow:
                        cell.fill = yellowFill
                        cell.font = darkYellowText

                    else:
                        # Default cell formatting
                        cell.font = Font(name="Calibri", size=10)

                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                starting_row += 1

            wb.save(file_to_save)

            return True
        except Exception as e:
            print(f"An error occurred: {e}")
            return False

    @staticmethod
    def to_xlsx_offside_noheader_color(
        file_to_save: str,
        ws_name: str,
        list_of_contents: list,
        list_of_red: list = [],
        list_of_green: list = [],
        list_of_yellow: list = [],
        col_offside: int = 0,
        starting_row: int = 1,
    ):
        try:
            # Check if workbook exists
            if os.path.exists(file_to_save):
                wb = load_workbook(file_to_save)
            else:
                wb = Workbook()
                wb.remove(wb["Sheet"])

            # Create sheet
            if ws_name in wb.sheetnames:
                ws_target = wb[ws_name]
            else:
                ws_target = wb.create_sheet(ws_name)

            for items in list_of_contents:
                for col, item in enumerate(items):
                    cell = ws_target.cell(
                        row=starting_row, column=col + 1 + col_offside
                    )
                    cell.value = item

                    # Apply formatting based on category
                    if item in list_of_red:
                        cell.fill = redFill
                        cell.font = darkRedText

                    elif item in list_of_green:
                        cell.fill = greenFill

                        cell.font = darkGreenText
                        # cell.font = Font(name="Calibri", size=10)
                    elif item in list_of_yellow:
                        cell.fill = yellowFill

                        cell.font = darkYellowText
                        # cell.font = Font(name="Calibri", size=10)
                    else:
                        # Default cell formatting
                        cell.font = Font(name="Calibri", size=10)

                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                starting_row += 1

            wb.save(file_to_save)

            return True
        except Exception as e:
            print(f"An error occurred: {e}")
            return False
